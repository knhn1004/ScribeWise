import cv2
import numpy as np
import os
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import yt_dlp
from transformers import pipeline


class VideoProcessor:

    def __init__(self, save_path, threshold=40.0):
        self.save_path = save_path
        self.threshold = threshold
        self.captioner = pipeline(
            "image-to-text", model="Salesforce/blip-image-captioning-base"
        )

        if not os.path.exists(save_path):
            os.makedirs(save_path)

    def download_video(self, url):
        ydl_opts = {
            "format": "bestvideo[height<=360][ext=mp4]+bestaudio/best[height<=360][ext=mp4]",
            "outtmpl": os.path.join(self.save_path, "%(title)s.%(ext)s"),
            "quiet": True,
            "no_warnings": True,
        }

        try:
            with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                info = ydl.extract_info(url, download=True)
                downloaded_file = os.path.join(self.save_path, f"{info['title']}.mp4")
                print(f"Downloaded '{info['title']}' successfully!")
                return downloaded_file
        except Exception as e:
            print(f"Error downloading video: {e}")
            return None

    def get_image_caption(self, image_path):
        result = self.captioner(image_path)
        return result[0]["generated_text"] if result else "No caption generated"

    def resize_image(self, image_path, max_size=(300, 300)):
        img = Image.open(image_path)
        img.thumbnail(max_size, Image.Resampling.LANCZOS)
        resized_path = image_path.replace(".png", "_resized.png")
        img.save(resized_path, "PNG")
        return resized_path

    @staticmethod
    def format_time(milliseconds):
        seconds, milliseconds = divmod(milliseconds, 1000)
        minutes, seconds = divmod(seconds, 60)
        hours, minutes = divmod(minutes, 60)
        return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}.{int(milliseconds):03}"

    def process_video(self, video_path):
        cap = cv2.VideoCapture(video_path)
        last_frame = None
        scene_count = 0
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Filename", "Scene Number", "Timestamp", "Caption"])

        while True:
            ret, frame = cap.read()
            if not ret:
                break

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

            if last_frame is not None:
                frame_diff = cv2.absdiff(last_frame, gray)

                if np.mean(frame_diff) > self.threshold:
                    timestamp = cap.get(cv2.CAP_PROP_POS_MSEC)
                    formatted_time = self.format_time(timestamp)
                    file_name = f"scene_{scene_count:04d}_{formatted_time}.png"
                    full_path = os.path.join(self.save_path, file_name)

                    if cv2.imwrite(full_path, frame):
                        print(f"Saved: {full_path}")
                        caption = self.get_image_caption(full_path)
                        resized_image_path = self.resize_image(full_path)
                        sheet.append([file_name, scene_count, formatted_time, caption])

                        img = OpenpyxlImage(resized_image_path)
                        img.anchor = f"E{sheet.max_row}"
                        sheet.add_image(img)

                        sheet.row_dimensions[sheet.max_row].height = 100
                        sheet.column_dimensions["E"].width = 40

                        scene_count += 1
                    else:
                        print(f"Failed to save: {full_path}")

            last_frame = gray

        cap.release()
        print(f"Detected {scene_count} scene changes.")

        excel_file_path = os.path.join(self.save_path, "scene_data_with_images.xlsx")
        workbook.save(excel_file_path)
        print(f"Excel file with images saved at {excel_file_path}")


def main():
    url = input("Enter the YouTube URL: ")
    save_path = "./"

    processor = VideoProcessor(save_path)

    video_path = processor.download_video(url)
    if video_path:
        processor.process_video(video_path)


if __name__ == "__main__":
    main()
